import sys
import os
import csv
import openpyxl

os.environ['PYTHONIOENCODING'] = 'utf-8'

def convert(xlsx_path, output_dir=None):
    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(xlsx_path))
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    wb = openpyxl.load_workbook(xlsx_path)
    sheets = {
        'Closed Positions': 'closed_positions.csv',
        'Dividends': 'dividends.csv',
        'Account Activity': 'account_activity.csv',
    }
    exported = []
    for sheet_name, csv_name in sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        out_path = os.path.join(output_dir, csv_name)
        with open(out_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                if all(v is None for v in row):
                    continue
                writer.writerow(['' if v is None else str(v) for v in row])
        exported.append(out_path)
    return exported

if __name__ == '__main__':
    if len(sys.argv) < 2:
        sys.exit(1)
    xlsx = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else None
    if not os.path.exists(xlsx):
        sys.exit(1)
    try:
        convert(xlsx, out)
    except Exception:
        sys.exit(1)